from openpyxl import load_workbook
EXCEL_FILE = "config/API_details.xlsx"

def load_api_config(file):
    
    get_apis = {}   # label → actual API name
    
    
    wb = load_workbook(file)
    sheet = wb.active

    headers = [cell.value for cell in sheet[1]]
    # print("++++++++++",headers)
    col_index = {name: idx for idx, name in enumerate(headers)}
    # print("++++++++++",col_index)
    config = {}
    # all_data_apis = {}  # label → actual API name
    all_data_apis = {}
    get_apis = {}
    current_api = None

    for row in sheet.iter_rows(min_row=2):
        
        # ✅ Add GET APIs automatically based on name
        name = row[col_index["API_NAME"]].value
        url = row[col_index["URL"]].value
        method = row[col_index["METHOD_TYPE"]].value
        req_cell = row[col_index["REQUEST_ARGUMENT_NAME"]]
        # all_data_value = row[col_index.get("AllDataAPI")].value
        if name and str(name).upper().startswith("GET"):
            get_apis[name] = name
        

        if url:
            current_api = {
                "Name": name,
                "URL": url,
                "Method Type": method,
                "Request Cells": []
            }

            config[name] = current_api

            # Map AllDataAPI label → actual API
            # if all_data_value:
            #     label = str(all_data_value).strip()
                # all_data_apis[label] = name

        if not current_api:
            continue

        if req_cell.value:
            lines = str(req_cell.value).split("\n")
            # print("FUFWUIEF",lines)
            for line in lines:
                line = line.strip()
                if not line:
                    continue

                current_api["Request Cells"].append({
                    "text": line,
                    "underline": req_cell.font.underline
                })

    # return config, all_data_apis
    # print("CONFIG",config["Save ISAC data"])
    return config, all_data_apis, get_apis

# API_CONFIG, ALL_DATA_APIS = load_api_config(EXCEL_FILE)
API_CONFIG, ALL_DATA_APIS, GET_APIS = load_api_config(EXCEL_FILE)
print("API_CONFIG",API_CONFIG)