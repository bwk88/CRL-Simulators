#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import tkinter as tk
from tkinter import ttk, scrolledtext, filedialog
import requests
import os
import json
import threading
from openpyxl import load_workbook

from ttkwidgets.autocomplete import AutocompleteCombobox

EXCEL_FILE = "API_details.xlsx"

# ---------------- LOAD CONFIG ----------------

def load_config():
    config_path = "config.json"

    if not os.path.exists(config_path):
        raise FileNotFoundError("config.json not found")

    try:
        with open(config_path) as f:
            return json.load(f)
    except Exception as e:
        print("Error loading config:", e)
        return {}

CONFIG = load_config()
BASE_URL = CONFIG.get("BASE_URL", "http://localhost:5000")


# ---------------- LOAD EXCEL ----------------

def load_api_config(file):
    
    get_apis = {}   # label → actual API name
    
    
    wb = load_workbook(file)
    sheet = wb.active

    headers = [cell.value for cell in sheet[1]]
    col_index = {name: idx for idx, name in enumerate(headers)}

    config = {}
    # all_data_apis = {}  # label → actual API name
    all_data_apis = {}
    get_apis = {}
    current_api = None

    for row in sheet.iter_rows(min_row=2):
        
        # ✅ Add GET APIs automatically based on name
        name = row[col_index["API_NAME"]].value
        url = row[col_index["URL"]].value
        method = row[col_index["METHOD_TYPE"]].value
        req_cell = row[col_index["REQUEST_ARGUMENT_NAME"]]
        # all_data_value = row[col_index.get("AllDataAPI")].value
        if name and str(name).upper().startswith("GET"):
            get_apis[name] = name
        

        if url:
            current_api = {
                "Name": name,
                "URL": url,
                "Method Type": method,
                "Request Cells": []
            }

            config[name] = current_api

            # Map AllDataAPI label → actual API
            # if all_data_value:
            #     label = str(all_data_value).strip()
                # all_data_apis[label] = name

        if not current_api:
            continue

        if req_cell.value:
            lines = str(req_cell.value).split("\n")

            for line in lines:
                line = line.strip()
                if not line:
                    continue

                current_api["Request Cells"].append({
                    "text": line,
                    "underline": req_cell.font.underline
                })

    # return config, all_data_apis
    return config, all_data_apis, get_apis


# API_CONFIG, ALL_DATA_APIS = load_api_config(EXCEL_FILE)
API_CONFIG, ALL_DATA_APIS, GET_APIS = load_api_config(EXCEL_FILE)

# ---------------- GLOBAL ----------------

dynamic_entries = {}

def api_call(url, payload):
    try:
        response = requests.post(url, json=payload)
        # print("API_Called: ", url)
        return {
            "status_code": response.status_code,
            "response": response.json()
        }
    except Exception as e:
        return {"status_code": "ERROR", "response": str(e)}

# ---------------- UI HELPERS ----------------

def clear_input_area():
    for widget in scrollable_frame.winfo_children():
        widget.destroy()


def generate_input_fields(cells):
    global dynamic_entries
    dynamic_entries = {}

    clear_input_area()

    if not cells:
        tk.Label(scrollable_frame, text="No input required").pack(anchor="w")
        return

    for item in cells:
        text = item["text"]
        is_heading = item["underline"] is not None

        if is_heading:
            tk.Label(
                scrollable_frame,
                text=text,
                font=("Arial", 10, "bold"),
                fg="blue"
            ).pack(anchor="w", pady=(10, 3))
        else:
            frame = tk.Frame(scrollable_frame)
            frame.pack(fill="x", pady=3)

            tk.Label(frame, text=text, width=25, anchor="w").pack(side="left")

            entry = tk.Entry(frame)
            entry.pack(side="left", fill="x", expand=True)

            dynamic_entries[text] = entry


def build_payload():
    selected_auto = all_api_dropdown.get()

    # 🔥 AUTO MODE
    if selected_auto:
        # api_name = ALL_DATA_APIS[selected_auto]
        api_name = GET_APIS[selected_auto]
        api_data = API_CONFIG[api_name]

        payload = {}
        for item in api_data["Request Cells"]:
            if item["underline"] is None:
                payload[item["text"]] = "0"

        return payload

    # 🔹 MANUAL MODE
    payload = {}
    for key, entry in dynamic_entries.items():
        payload[key] = entry.get().strip()

    return payload


# ---------------- EVENTS ----------------

def on_api_select(event=None):
    all_api_dropdown.set("")  # reset auto dropdown

    selected = api_dropdown.get()
    api_data = API_CONFIG[selected]

    generate_input_fields(api_data["Request Cells"])


def on_all_api_select(event=None):
    api_dropdown.set("")  # reset manual dropdown

    selected_label = all_api_dropdown.get()
    # api_name = ALL_DATA_APIS[selected_label]
    api_name = GET_APIS[selected_label]
    clear_input_area()

    tk.Label(
        scrollable_frame,
        text=f"⚡ Auto Mode: {api_name} (all values = 0)",
        fg="green",
        font=("Arial", 10, "bold")
    ).pack(anchor="w", pady=10)

    call_api_thread()  # auto call


def call_api():
    selected_auto = all_api_dropdown.get()

    if selected_auto:
        # selected = ALL_DATA_APIS[selected_auto]
        selected = GET_APIS[selected_auto]
    else:
        selected = api_dropdown.get()

    api_data = API_CONFIG[selected]

    url = f"{BASE_URL}{api_data['URL']}"
    print("API_Called: ", url)
    
    method = str(api_data["Method Type"]).upper()
    

    payload = build_payload()

    if method == "DELETE":
        result = call_api_delete(url, payload)
    else:
        result = api_call(url, payload)

    root.after(0, update_ui, result, method)


def call_api_delete(url, payload):
    try:
        response = requests.delete(url, json=payload)

        try:
            res_json = response.json()
        except:
            res_json = {"raw_response": response.text}

        return {
            "status_code": response.status_code,
            "response": res_json
        }

    except Exception as e:
        return {"status_code": "ERROR", "response": str(e)}


def update_ui(result, method):
    status = result["status_code"]
    response = result["response"]

    message = ""

    if isinstance(response, dict):

        if method == "POST":

            if status == 200:
                flag = response.get("success_failure_flag")

                if flag == 1:
                    message = "✅ Success"
                else:
                    remarks = response.get("remarks", "Unknown error")
                    message = f"❌ Failed: {remarks}"

            elif status in [400, 404, 500]:
                reason = response.get("reason", "Unknown error")
                message = f"❌ Error {status}: {reason}"

            else:
                message = f"⚠️ Unexpected Status: {status}"

        else:
            message = f"✅ GET Success (Status {status})"

    else:
        message = str(response)

    color = "green" if "✅" in message else "red"

    status_label.config(text=message, fg=color)

    result_box.delete(1.0, tk.END)
    result_box.insert(tk.END, json.dumps(result, indent=4))


def call_api_thread():
    status_label.config(text="⏳ Calling API...", fg="black")
    threading.Thread(target=call_api, daemon=True).start()


# ---------------- FILE OPS ----------------

def load_json():
    path = filedialog.askopenfilename(filetypes=[("JSON", "*.json")])
    if not path:
        return

    with open(path) as f:
        data = json.load(f)

    result_box.delete(1.0, tk.END)
    result_box.insert(tk.END, json.dumps(data, indent=4))


def save_response():
    content = result_box.get(1.0, tk.END).strip()
    if not content:
        return

    path = filedialog.asksaveasfilename(defaultextension=".json")
    if path:
        with open(path, "w") as f:
            f.write(content)


# ---------------- UI SETUP ----------------

root = tk.Tk()
root.title("API Client Tool")
root.geometry("950x720")
root.configure(bg="#f4f6f9")

# HEADER
header = tk.Frame(root, bg="#1f2a44", height=60)
header.pack(fill="x")

tk.Label(header, text="🚀 API Client Dashboard", bg="#1f2a44",
         fg="white", font=("Segoe UI", 16, "bold")).pack(pady=10)

main = tk.Frame(root, bg="#f4f6f9")
main.pack(fill="both", expand=True, padx=10, pady=10)

# DROPDOWNS
card1 = tk.Frame(main, bg="white", bd=1, relief="solid")
card1.pack(fill="x", pady=5)

# Manual
top_frame = tk.Frame(card1, bg="white")
top_frame.pack(anchor="w", padx=10, pady=10)

tk.Label(top_frame, text="Select API:", bg="white",
         font=("Segoe UI", 10, "bold")).pack(side="left")

api_dropdown = AutocompleteCombobox(top_frame,
                            completevalues=list(API_CONFIG.keys()),
                            width=45)
api_dropdown.pack(side="left", padx=10)
api_dropdown.bind("<<ComboboxSelected>>", on_api_select)

# Auto
all_frame = tk.Frame(card1, bg="white")
all_frame.pack(anchor="w", padx=10, pady=(0, 10))

tk.Label(all_frame, text="All Data APIs:", bg="white",
         font=("Segoe UI", 10, "bold")).pack(side="left")

# all_api_dropdown = ttk.Combobox(all_frame,
#                                 values=list(ALL_DATA_APIS.keys()),
#                                 state="readonly", width=45)

all_api_dropdown = ttk.Combobox(all_frame,
                                values=list(GET_APIS.keys()),
                                state="readonly", width=45)


all_api_dropdown.pack(side="left", padx=10)
all_api_dropdown.bind("<<ComboboxSelected>>", on_all_api_select)

# INPUT AREA
input_card = tk.Frame(main, bg="white", bd=1, relief="solid")
input_card.pack(fill="both", pady=10)

tk.Label(input_card, text="🧾 Input Parameters", bg="white",
         font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=10, pady=5)

canvas = tk.Canvas(input_card, height=200, bg="white", highlightthickness=0)
scrollbar = tk.Scrollbar(input_card, orient="vertical", command=canvas.yview)

scrollable_frame = tk.Frame(canvas, bg="white")
scrollable_frame.bind("<Configure>",
    lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
canvas.configure(yscrollcommand=scrollbar.set)

canvas.pack(side="left", fill="both", expand=True, padx=10, pady=5)
scrollbar.pack(side="right", fill="y")

# BUTTONS
btn_frame = tk.Frame(main, bg="#f4f6f9")
btn_frame.pack(pady=10)

tk.Button(btn_frame, text="▶ Call API", command=call_api_thread,
          bg="#007bff", fg="white").pack(side="left", padx=10)

tk.Button(btn_frame, text="📂 Load JSON", command=load_json,
          bg="#6c757d", fg="white").pack(side="left", padx=10)

tk.Button(btn_frame, text="💾 Save Response", command=save_response,
          bg="#28a745", fg="white").pack(side="left", padx=10)

# STATUS
status_label = tk.Label(main, text="", bg="#f4f6f9",
                        font=("Segoe UI", 10, "bold"))
status_label.pack(pady=5)

# RESPONSE
result_box = scrolledtext.ScrolledText(main, font=("Consolas", 10))
result_box.pack(fill="both", expand=True, padx=10, pady=5)

# INIT
if API_CONFIG:
    api_dropdown.current(0)
    on_api_select()

root.mainloop()
